import os
import csv as stdcsv
import pandas as pd
import numpy as np
import glob
import re
import sys
import html
import gc
from tqdm import tqdm
import joblib

# NLTK e spaCy per elaborazioni testuali
import nltk
from nltk.corpus import stopwords
import spacy
import spacy.cli

# Per TF-IDF
from sklearn.feature_extraction.text import TfidfVectorizer

# SBERT
from sentence_transformers import SentenceTransformer

# ---------------- Directory e file ----------------
ROOT_DIR       = r"D:\Magistrale\AI"
CSV_DIR        = os.path.join(ROOT_DIR, "output_chunks")
CLUSTER_DIR    = os.path.join(ROOT_DIR, "output_cluster")
EXCEL_PATH     = os.path.join(ROOT_DIR, "tickets_features_ngrams_tfidf_sbert.xlsx")
CSV_FALLBACK   = os.path.join(CSV_DIR, "fattoAmano.csv")
NPZ_PATH       = os.path.join(ROOT_DIR, "features.npz")
VEC_PATH       = os.path.join(CLUSTER_DIR, "tfidf_vectorizer.joblib")

# Creazione delle cartelle, se non esistono
os.makedirs(CSV_DIR, exist_ok=True)
os.makedirs(CLUSTER_DIR, exist_ok=True)
os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)

# ---------------- Parametri ML ----------------
SBERT_MODEL     = "sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2"
BATCH_SIZE      = 32
SAMPLE_SIZE     = 5000
TEMPORAL_WEIGHT = 0.0
TITLE_WEIGHT    = 1.2
TFIDF_WEIGHT    = 1.0
default_tfidf_params = {'token_pattern': r"[^\s]+", 'min_df': 1}

# Aumento limite per campi CSV
try:
    stdcsv.field_size_limit(2**31 - 1)
except OverflowError:
    stdcsv.field_size_limit(sys.maxsize)

# ---------------- Setup NLP ----------------
nltk.download('stopwords', quiet=True)
italian_stops = set(stopwords.words('italian'))
try:
    nlp = spacy.load('it_core_news_sm')
except OSError:
    spacy.cli.download('it_core_news_sm')
    nlp = spacy.load('it_core_news_sm')

# ---------------- Funzioni utili ----------------

def detect_delimiter(path):
    with open(path, 'r', encoding='utf-8', errors='ignore') as f:
        sample = f.read(2048)
    try:
        return stdcsv.Sniffer().sniff(sample).delimiter
    except stdcsv.Error:
        return ','


def safe_read_csv(path, **kwargs):
    try:
        return pd.read_csv(path, encoding='utf-8', engine='python', on_bad_lines='skip', **kwargs)
    except:
        return pd.read_csv(path, encoding='latin-1', engine='python', on_bad_lines='skip', **kwargs)


def clean_html(raw: str) -> str:
    if pd.isna(raw): return ""
    text = html.unescape(raw)
    text = re.sub(r"<[^>]+>", " ", text)
    return ' '.join(text.split()).lower()


def custom_tokenizer(text: str) -> list:
    doc = nlp(text)
    toks = []
    for t in doc:
        lm = t.lemma_.lower()
        if t.is_alpha and len(lm) > 2 and lm not in italian_stops and not t.is_stop:
            toks.append(lm)
    return toks


def generate_ngrams(tokens: list) -> list:
    ng = tokens.copy()
    L = len(tokens)
    if L >= 2:
        for i in range(L-1):
            ng.append(tokens[i] + "_" + tokens[i+1])
    if L >= 3:
        for i in range(L-2):
            ng.append(tokens[i] + "_" + tokens[i+1] + "_" + tokens[i+2])
    return ng


def char_ngrams(text: str, min_n=4, max_n=5) -> list:
    c = text.replace(' ', '_')
    ng = []
    L = len(c)
    for n in range(min_n, max_n+1):
        for i in range(L-n+1):
            ng.append(c[i:i+n])
    return ng


def cyclic_feats(vals, max_val):
    theta = 2 * np.pi * vals / max_val
    return np.sin(theta), np.cos(theta)

# ---------------- Batch iniziale ----------------
if __name__ == "__main__":
    ticket_dfs = []
    files = [p for p in glob.glob(os.path.join(CSV_DIR, "*")) if os.path.isfile(p) and p != CSV_FALLBACK]
    if not files and os.path.exists(CSV_FALLBACK):
        files = [CSV_FALLBACK]

    for path in tqdm(files, desc="Lettura file"):
        try:
            sep = detect_delimiter(path)
            head = safe_read_csv(path, sep=sep, nrows=5)
            cols = head.columns.tolist()
            date_cols = [c for c in cols if 'create' in c.lower()]
            use_cols  = [c for c in cols if any(x in c.lower() for x in ['segna','title','descr','create'])]
            df = safe_read_csv(path, sep=sep, usecols=use_cols, parse_dates=date_cols)
        except Exception as e:
            print(f"Warning salto {path}: {e}")
            continue

        df.rename(columns={c: (
            'ticket_id'  if 'segna' in c.lower() else
            'title'      if 'title' in c.lower() else
            'descr'      if 'descr' in c.lower() else
            'created_at' if 'create' in c.lower() else c
        ) for c in df.columns}, inplace=True)

        df['title'] = df.get('title', '').fillna('').astype(str)
        df['descr'] = df.get('descr', '').fillna('').astype(str)
        df['title_clean'] = df['title'].apply(clean_html)
        df['descr_clean'] = df['descr'].apply(clean_html)
        df['title_tokens'] = df['title_clean'].apply(custom_tokenizer)
        df['descr_tokens'] = df['descr_clean'].apply(custom_tokenizer)
        df['title_ngrams_combined'] = df.apply(lambda r: ' '.join(generate_ngrams(r['title_tokens']) + char_ngrams(r['title_clean'])), axis=1)
        df['descr_ngrams_combined'] = df.apply(lambda r: ' '.join(generate_ngrams(r['descr_tokens']) + char_ngrams(r['descr_clean'])), axis=1)

        ticket_dfs.append(df[['ticket_id','title_ngrams_combined','descr_ngrams_combined','created_at']])

    _df = pd.concat(ticket_dfs, ignore_index=True) if ticket_dfs else pd.DataFrame(columns=['ticket_id','title_ngrams_combined','descr_ngrams_combined','created_at'])
    _df['created_at'] = pd.to_datetime(_df['created_at'], errors='coerce')
    limit = pd.to_datetime('today') - pd.Timedelta(days=365)
    _df = _df[_df['created_at'] >= limit]
    if SAMPLE_SIZE and len(_df) > SAMPLE_SIZE:
        _df = _df.sample(SAMPLE_SIZE, random_state=42).reset_index(drop=True)

    texts = (_df['title_ngrams_combined'] + ' ' + _df['descr_ngrams_combined']).tolist()
    tfidf_vec = TfidfVectorizer(**default_tfidf_params)
    X_tfidf = tfidf_vec.fit_transform(texts).toarray()
    norms = np.linalg.norm(X_tfidf, axis=1, keepdims=True)
    norms[norms == 0] = 1
    X_tfidf /= norms
    joblib.dump(tfidf_vec, VEC_PATH)

    model = SentenceTransformer(SBERT_MODEL)
    X_sbert = model.encode(texts, batch_size=BATCH_SIZE, convert_to_numpy=True, normalize_embeddings=True)

    X_comb = np.hstack([TITLE_WEIGHT * X_sbert, TFIDF_WEIGHT * X_tfidf])
    X_all  = np.hstack([X_comb, np.zeros((X_comb.shape[0],4))])
    mask_nonzero = np.any(X_all != 0, axis=1)
    X_all = X_all[mask_nonzero]
    _df   = _df[mask_nonzero].reset_index(drop=True)

    np.savez_compressed(NPZ_PATH, X=X_all)
    _df.to_parquet(os.path.join(ROOT_DIR,'tickets_features_ngrams_tfidf_sbert.parquet'), index=False)
    print('Batch completato')

# ---------------- Classe TicketEmbedder ----------------
class TicketEmbedder:
    def __init__(self):
        if not os.path.exists(VEC_PATH):
            raise FileNotFoundError('Esegui batch iniziale con Load_Tickets.py prima di usare embed()')
        self.tfidf_vec = joblib.load(VEC_PATH)
        data = np.load(NPZ_PATH)
        self.dim = data['X'].shape[1]
        self.model = SentenceTransformer(SBERT_MODEL)
        os.makedirs(os.path.dirname(CSV_FALLBACK), exist_ok=True)
        new_file = not os.path.exists(CSV_FALLBACK)
        self.csv_file = open(CSV_FALLBACK,'a',newline='',encoding='utf-8-sig')
        fieldnames=['ticket_id','title','descr','created_at']
        self.csv_writer=stdcsv.DictWriter(self.csv_file,fieldnames=fieldnames,delimiter=';',quotechar='"',quoting=stdcsv.QUOTE_MINIMAL,escapechar='\\')
        if new_file:
            self.csv_writer.writeheader()

    def embed(self, ticket_id, title, descr, created_at):
        if isinstance(created_at, str):
            created_at = pd.to_datetime(created_at, errors='coerce')
        tc = clean_html(title)
        dc = clean_html(descr)
        tt = custom_tokenizer(tc)
        dt = custom_tokenizer(dc)
        comb = ' '.join(generate_ngrams(tt) + char_ngrams(tc) + generate_ngrams(dt) + char_ngrams(dc))

        # TF-IDF
        Xt = self.tfidf_vec.transform([comb]).toarray()
        norms = np.linalg.norm(Xt, axis=1, keepdims=True)
        norms[norms == 0] = 1
        Xt /= norms

        # SBERT
        Xs = self.model.encode([comb], batch_size=BATCH_SIZE, convert_to_numpy=True, normalize_embeddings=True)
        feat = np.hstack([TITLE_WEIGHT * Xs, TFIDF_WEIGHT * Xt]).flatten()

        # Temporali
        hs, hc = cyclic_feats(np.array([created_at.hour]), 24)
        ws, wc = cyclic_feats(np.array([created_at.weekday()]), 7)
        temp = np.hstack([hs, hc, ws, wc]) * TEMPORAL_WEIGHT
        Xnew = np.hstack([feat, temp])

        if Xnew.size != self.dim:
            raise ValueError(f"Embedding dimension {Xnew.size} != {self.dim}, ricrea batch iniziale.")

        # Append memmap
        npy_path = NPZ_PATH.replace('.npz', '.npy')
        def append_memmap(path, Xnew):
            if not os.path.exists(path):
                mmap = np.lib.format.open_memmap(path, mode='w+', dtype=Xnew.dtype, shape=(1, Xnew.shape[0]))
                mmap[0] = Xnew; mmap.flush(); mmap._mmap.close()
            else:
                old = np.lib.format.open_memmap(path, mode='r+', dtype=Xnew.dtype)
                n_old, dim = old.shape
                tmp_path = path + '.tmp'
                mmap_new = np.lib.format.open_memmap(tmp_path, mode='w+', dtype=old.dtype, shape=(n_old+1, dim))
                mmap_new[:n_old] = old[:]
                mmap_new[n_old] = Xnew
                mmap_new.flush()
                old._mmap.close(); mmap_new._mmap.close(); del old, mmap_new; gc.collect(); os.replace(tmp_path, path)
        append_memmap(npy_path, Xnew)
        np.savez_compressed(NPZ_PATH, X=np.lib.format.open_memmap(npy_path, mode='r', dtype=Xnew.dtype))

        # Scrittura risultati
        created_at_str = created_at.strftime('%Y-%m-%d %H:%M:%S.%f')
        row = {'ticket_id': ticket_id, 'title': title, 'descr': descr, 'created_at': created_at_str}
        from openpyxl import load_workbook, Workbook
        try:
            wb = load_workbook(EXCEL_PATH); ws = wb.active; ws.append([ticket_id, title, descr, created_at_str])
        except FileNotFoundError:
            wb = Workbook(); ws = wb.active; ws.append(['ticket_id','title','descr','created_at']); ws.append([ticket_id, title, descr, created_at_str])
        wb.save(EXCEL_PATH)
        self.csv_writer.writerow(row); self.csv_file.flush()

        return ticket_id, Xnew
